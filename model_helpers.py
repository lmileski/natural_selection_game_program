"""
Helper functions for the model component of the program
"""

import os
from platform import system
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import pandas as pd
import subprocess

def produce_diagonal_matrix(rectangular_matrix: list[list[int]]) -> list[list[int]]:
    """
    Goes through a rectangular matrix (each inner list contains an int for a square in a column)
    Diagonal matrix will be needed for the ui's wipe transition from top left square to bottom right square
    Each inner list contains a diagonal section of the board - in a 3x3 matrix the change looks like this:
    [0] [1] [2]      [0] [1] [2]
    [0] [1] [2]  ->  [1] [2] [3]
    [0] [1] [2]      [2] [3] [4]
    where each integer corresponds to the index of that item's inner list

    Preconditions:
        - length and width are equal
        - 2 <= side length <= 8

    Doctests:
    >>> produce_diagonal_matrix([[0, 1, 2], [1, 2, 3], [2, 3, 4]])
    [[0], [1, 1], [2, 2, 2], [3, 3], [4]]
    >>> produce_diagonal_matrix([[0, 1, 2, 3, 4], [1, 2, 3, 4, 5], [2, 3, 4, 5 ,6], [3, 4, 5, 6, 7], [4, 5, 6, 7, 8]])
    [[0], [1, 1], [2, 2, 2], [3, 3, 3, 3], [4, 4, 4, 4, 4], [5, 5, 5, 5], [6, 6, 6], [7, 7], [8]]
    """
    # checking precondition that rectangular matrix is actually rectangular
    num_columns = len(rectangular_matrix)
    # side lengths must be between 2-8, inclusive
    assert 2 <= num_columns <= 8
    for column in rectangular_matrix:
        assert num_columns == len(column)
    
    board_length = len(rectangular_matrix) # height and width must be equal
    num_inner_lists = board_length*2-1
    diagonal_matrix: list[list[int]] = [[] for i in range(num_inner_lists)]
    row = 0
    for x, column in enumerate(rectangular_matrix):
        inner_list_i = x # indexed list always starts as the column number
        while row < board_length:
            diagonal_matrix[inner_list_i].append(column[row])
            row += 1
            inner_list_i += 1
        row = 0

    return diagonal_matrix

def determine_round_winner(array_of_outcomes: list[list[int]]) -> int:
    """
    Determines whether predators or prey won the round.
    Winner is based off which animal team won a majority of squares
    The winner of each square is found in the array_of_outcomes argument

    Function is to be called using the returned array_of_outcomes from
    the Board class's modify_board_survivors() method

    Parameters:
        - array_of_outcomes (list[list[int]]): a 2d array with each integer representing the winner of a square
            Outer list holds every column, each inner list is a column that holds square outcomes for top to bottom
            If the int is -1, the predators won the square
            If the int is 1, the prey won the square
            If the int is 0, it was a tie
    
    Returns:
        - (int): 2 if the predators won, 3 if the prey won, and 4 if it was a tie

    Doctests:
    >>> determine_round_winner([[1, -1], [1, 0]])
    3
    >>> determine_round_winner([[1, -1, -1], [0, 1, -1], [0, 0, -1]])
    2
    >>> determine_round_winner([[0]])
    4
    """
    # keeps track of which team is currently winning the round
    winner_tracker = 0
    # adding points based of each squares winner
    for column in array_of_outcomes:
        for outcome in column:
            winner_tracker += outcome
    # positive result -> prey win
    if winner_tracker > 0:
        return 3
    # negative result -> predators win
    elif winner_tracker < 0:
        return 2
    # score of 0 -> tie
    else:
        return 4

def record_round_data(round_data: tuple[int, tuple[int, int], tuple[float, float], str], round_log_file='game_results_logs/round_log.csv') -> None:
        """
        Appends all relevant data to game_results_logs/round_log.csv using the pandas library
        Function to be used at the beginning of the game and end of every round
        Parameters:
            - round_data (tuple[int, tuple[int, int], tuple[float, float]]):
                - (0) current round
                - (1) total populations of predators and prey
                - (2) average skill levels of predators and prey
                - (3) animal team that won the round (either 'predators', 'prey', 'tie', or 'n/a')
            
            - round_log_file='game_results_logs/round_log.csv': the filename to record the round's results
        """
        # recording round's data
        df = pd.DataFrame({
            'Round': [round_data[0]], # starting data is marked as round 0, end of round 1 is marked as round 1, etc.
            'Predator Population': [round_data[1][0]],
            'Prey Population': [round_data[1][1]],
            'Average Predator Level': [round_data[2][0]],
            'Average Prey Level': [round_data[2][1]],
            'Animal Team Winner': [round_data[3]]
        })

        # only adding column headings on first round
        if round_data[0] == 0:
            open(round_log_file, 'w').close() # clearing file if its the first round
            heading = True
        else:
            heading = False
        # appending this round's data to a csv file - will be used if user wants to export results to an excel file
        df.to_csv(round_log_file, mode='a', index=False, header=heading)

def record_start_and_end_data(skill_levels_to_populations: tuple[dict[int, int], dict[int, int]], start_of_game: bool) -> None:
    """
    Records the predator and prey populations with respect to each's skill levels
    Function only to be called at the start and end of game

    Parameters:
        - skill_levels_to_populations (tuple[dict[int, int], dict[int, int]]):
            - (0) list of predators skill levels to population - inside each tuple in the list contains:
                - (0) predator skill level
                - (1) predator population at skill level
            - (1) list of prey skill levels to population - inside each tuple in the list contains:
                - (0) prey skill level
                - (1) prey population at skill level
        
        - start_of_game (bool): True if start of game, False if end of game
    """
    # dictionary holding predator and prey skill levels to population size data
    data = {}

    # finding the highest animal's level and using that value for the range of level-population rows
    highest_animal_level = 0
    for predator_level in skill_levels_to_populations[0]: # first looping over predator levels
        assert predator_level >= 0, "Level must be a non-negative integer"
        if predator_level > highest_animal_level:
            highest_animal_level = predator_level
    
    for prey_level in skill_levels_to_populations[1]: # then looping over prey levels
        assert prey_level >= 0, "Level must be a non-negative integer"
        if prey_level > highest_animal_level:
            highest_animal_level = prey_level

    # changes preceding word for column header depending of whether its the start or end of game data
    if start_of_game:
        preceding_word = 'Starting '
    else:
        preceding_word = 'Ending '

    # recording levels to population data only if passed (start and end of game)
    for i, dict_of_levels_to_population in enumerate(skill_levels_to_populations): # (i=0): predators, (i=1): prey
        if i == 0: # recording predator populations
            # key is the column header and value is a list of levels from 0 to the highest animal's level
            data[preceding_word + 'Predator Level'] = [x for x in range(highest_animal_level+1)]
            data[preceding_word + 'Respective Predator Population'] = [0 for _ in range(highest_animal_level+1)]
            for level, population in dict_of_levels_to_population.items():
                # adding the population of predators at their respective skill level
                data[preceding_word + 'Respective Predator Population'][level] = population
        else: # recording prey populations
            # key is the column header and value is a list of levels from 0 to the highest animal's level
            data[preceding_word + 'Prey Level'] = [x for x in range(highest_animal_level+1)]
            data[preceding_word + 'Respective Prey Population'] = [0 for _ in range(highest_animal_level+1)]
            for level, population in dict_of_levels_to_population.items():
                # adding the population of predators at their respective skill level
                data[preceding_word + 'Respective Prey Population'][level] = population

    # appending this round's data to a csv file - will be used if user wants to export results to an excel file
    df = pd.DataFrame(data)
    if start_of_game:
        df.to_csv('game_results_logs/start_of_game_log.csv', mode='w', index=False)
    else:
        df.to_csv('game_results_logs/end_of_game_log.csv', mode='w', index=False)

def transfer_round_logs(completed_game_round_log = 'game_results_logs/round_log.csv',
                        previous_game_round_log = 'game_results_logs/last_game_round_log.csv') -> None:
    """
    Transfers the current completed game's log of round data to the previous game's
    round log file
    """
    with open(completed_game_round_log) as f_from, open(previous_game_round_log, 'w', newline='') as f_to:
        data = f_from.read()
        f_to.write(data)

def get_new_filename(filename):
    """
    Finds a proper suffix for the excel file name depending on number of
    excel files already generated from loading previous games' data.
    """
    if not os.path.isfile(filename):
        return filename
    else:
        base, extension = os.path.splitext(filename)
        i = 1
        while os.path.isfile(f"{base}({i}){extension}"):
            i += 1
        return f"{base}({i}){extension}"

def export_game_data_to_excel(round_data_filename='game_results_logs/last_game_round_log.csv',
                              start_of_game_data_filename='game_results_logs/start_of_game_log.csv',
                              end_of_game_data_filename='game_results_logs/end_of_game_log.csv',
                              game_configurations_filename='game_settings_logs/user_configurations.json') -> None:
    """
    Creates an new excel file, opens it for the user, and uses the game's result data to
    generate tables, charts, and graphs, all with the pandas library.

    Parameters:
        - round_data_filename='game_results_logs/round_log.csv': the round data (predator/prey populations and avg. levels)
        - start_of_game_data_filename='game_results_logs/start_of_game_log.csv': the starting game data (predator/prey skill levels-populations)
        - end_of_game_data_filename='game_results_logs/end_of_game_log.csv': the ending game data (predator/prey skill levels-populations)
        - game_configurations_filename='game_settings_logs/previous_game_configurations.json': the last game's configurations
    """
    # creating df's for each file
    round_df = pd.read_csv(round_data_filename)
    start_of_game_df = pd.read_csv(start_of_game_data_filename)
    end_of_game_df = pd.read_csv(end_of_game_data_filename)
    game_configurations_df = pd.read_json(game_configurations_filename)

    # finding proper filename
    filename = get_new_filename('Natural_Selection_Game_Data.xlsx')

    # writing the tables to the new excel file
    with pd.ExcelWriter(filename) as writer:
        round_df.to_excel(writer, index=False, sheet_name='Round Data')
        start_of_game_df.to_excel(writer, index=False, sheet_name='Start of Game')
        end_of_game_df.to_excel(writer, index=False, sheet_name='End of Game')
        game_configurations_df.to_excel(writer, index=False, sheet_name='Game Configurations')

    # loading the workbook and grabbing a list of all sheets
    wb = load_workbook(filename)
    sheets = wb.worksheets

    # for each new sheet, adjusting the width of every column to fit the the text's max width 
    for sheet in sheets:
        # walking through columns
        for column in sheet.columns:
            max_length = 0
            # column = list(column) # type: ignore
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        # sheet.columns includes only cells with values
                        max_length = len(cell.value) # type: ignore
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    # Save the workbook
    wb.save(filename)
    # open the workbook
    open_file(filename)

    # export_game_data_to_excel() 

def open_file(path: str):
    """
    Loads a file based off the user's operating system
    """
    os_type = system()
    
    if os_type == 'Windows':
        os.startfile(path)
    elif os_type == 'Darwin':  # macOS
        subprocess.run(["open", path], check=True)
    elif os_type == 'Linux':
        subprocess.run(["xdg-open", path], check=True)
    # doesn't open the file for user if not windows, mac, or linux os